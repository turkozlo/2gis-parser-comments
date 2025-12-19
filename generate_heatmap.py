import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def generate_heatmap(df=None):
    csv_file = 'sberbank_all_reviews.csv'
    output_file = 'heatmap_report.xlsx'
    
    if df is None:
        if not os.path.exists(csv_file):
            print(f"Error: {csv_file} not found.")
            return None
            
        print(f"Loading {csv_file}...")
        df = pd.read_csv(csv_file)
        
        # Convert Date to datetime
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
        
        # Filter for 2025 (Jan 1 - Dec 10)
        df = df[
            (df['Date'] >= '2025-01-01') &
            (df['Date'] <= '2025-12-10')
        ]
        print(f"Found {len(df)} reviews for 2025 (Jan 1 - Dec 10)")
    else:
        print(f"Generating heatmap for provided DataFrame ({len(df)} reviews)...")

    
    # Filter negative reviews (exclude positive/no comment)
    df_negative = df[df['Tags'] != '#Позитивное/без_комментария'].copy()
    
    # Filter out rows with missing data or undefined GOSB
    df_negative = df_negative.dropna(subset=['Tags', 'Sub_tag', 'GOSB', 'City', 'Address'])
    df_negative = df_negative[
        (df_negative['Tags'] != '') & 
        (df_negative['Sub_tag'] != '') & 
        (df_negative['GOSB'] != '') & 
        (df_negative['GOSB'] != 'Не определено') &  # Exclude undefined GOSB
        (df_negative['City'] != '') & 
        (df_negative['Address'] != '')
    ]
    
    print(f"Processing {len(df_negative)} negative reviews...")
    
    # Split tags (in case there are multiple)
    # For simplicity, we'll use the first tag
    df_negative['Primary_Tag'] = df_negative['Tags'].apply(lambda x: x.split(',')[0].strip())
    
    # Exclude "Прочее" from tags and sub-topics
    df_negative = df_negative[
        (df_negative['Primary_Tag'] != '#Прочее') &
        (df_negative['Sub_tag'] != 'Прочее')
    ]
    
    # Combine City and Address for display
    df_negative['City_Address'] = df_negative['City'] + ', ' + df_negative['Address']
    
    # Create pivot table
    pivot = df_negative.groupby(
        ['Primary_Tag', 'Sub_tag', 'GOSB', 'City_Address']
    ).size().reset_index(name='Count')
    
    # Create a multi-level structure
    print("Building hierarchical table...")
    
    # Get unique combinations
    # SWAP: Now tags_subtags will be columns, gosb_addresses will be rows
    tags_subtags = pivot[['Primary_Tag', 'Sub_tag']].drop_duplicates().sort_values(['Primary_Tag', 'Sub_tag'])
    gosb_addresses = pivot[['GOSB', 'City_Address']].drop_duplicates().sort_values(['GOSB', 'City_Address'])
    
    # Create the data matrix
    # Rows: GOSB + City_Address combinations (swapped)
    # Columns: Tag + Sub_tag combinations (swapped)
    
    print("Creating Excel workbook...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Heatmap"
    
    # Build column headers (Tag -> Sub_tag) - now on top
    col_offset = 2  # Columns A and B for GOSB and City_Address
    current_col = col_offset + 1
    
    # Track Tag spans for merging
    tag_spans = {}
    current_tag = None
    start_col = current_col
    
    for idx, (tag, sub_tag) in enumerate(tags_subtags.values):
        # Row 1: Tag (will be merged later)
        if tag != current_tag:
            if current_tag is not None:
                tag_spans[current_tag] = (start_col, current_col - 1)
            current_tag = tag
            start_col = current_col
        
        # Row 2: Sub_tag
        ws.cell(row=2, column=current_col, value=sub_tag)
        current_col += 1
    
    # Add last tag span
    if current_tag is not None:
        tag_spans[current_tag] = (start_col, current_col - 1)
    
    # Write Tag headers and merge
    for tag, (start, end) in tag_spans.items():
        ws.cell(row=1, column=start, value=tag)
        if start < end:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
    
    # Set headers for GOSB and City_Address columns
    ws.cell(row=1, column=1, value="ГОСБ")
    ws.cell(row=1, column=2, value="Отделение")
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    
    # Build row data (GOSB + City_Address)
    current_row = 3
    gosb_row_spans = {}
    current_gosb = None
    start_row = current_row
    
    for gosb, city_address in gosb_addresses.values:
        # Column A: GOSB
        if gosb != current_gosb:
            if current_gosb is not None:
                gosb_row_spans[current_gosb] = (start_row, current_row - 1)
            current_gosb = gosb
            start_row = current_row
        
        # Column B: City_Address
        ws.cell(row=current_row, column=2, value=city_address)
        
        # Fill counts for this row
        for col_idx, (tag, sub_tag) in enumerate(tags_subtags.values):
            count = pivot[
                (pivot['Primary_Tag'] == tag) &
                (pivot['Sub_tag'] == sub_tag) &
                (pivot['GOSB'] == gosb) &
                (pivot['City_Address'] == city_address)
            ]['Count'].sum()
            
            if count > 0:
                ws.cell(row=current_row, column=col_offset + 1 + col_idx, value=int(count))
        
        current_row += 1
    
    # Add last GOSB span
    if current_gosb is not None:
        gosb_row_spans[current_gosb] = (start_row, current_row - 1)
    
    # Write GOSB column and merge
    for gosb, (start, end) in gosb_row_spans.items():
        ws.cell(row=start, column=1, value=gosb)
        if start < end:
            ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
    
    # Apply formatting
    print("Applying formatting...")
    
    # Header formatting
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=current_col-1):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Apply borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=current_row-1, min_col=1, max_col=current_col-1):
        for cell in row:
            cell.border = thin_border
            if cell.row >= 3:  # Data rows
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Conditional formatting for heatmap (color scale)
    from openpyxl.formatting.rule import ColorScaleRule
    
    # Find max value for color scale
    max_val = 0
    for row in ws.iter_rows(min_row=3, max_row=current_row-1, min_col=col_offset+1, max_col=current_col-1):
        for cell in row:
            if cell.value and isinstance(cell.value, (int, float)):
                max_val = max(max_val, cell.value)
    
    if max_val > 0:
        color_scale = ColorScaleRule(
            start_type='num', start_value=0, start_color='FFFFFF',
            mid_type='num', mid_value=max_val/2, mid_color='FFD966',
            end_type='num', end_value=max_val, end_color='FF0000'
        )
        
        # Apply to data range
        data_range = f"{get_column_letter(col_offset+1)}3:{get_column_letter(current_col-1)}{current_row-1}"
        ws.conditional_formatting.add(data_range, color_scale)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    for col_idx in range(col_offset + 1, current_col):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    
    # Generate analysis sheet
    print("Generating problem analysis...")
    generate_analysis_sheet(wb, pivot, df_negative)
    
    # Save
    wb.save(output_file)
    print(f"✓ Heatmap saved to {output_file}")
    print(f"  Total rows: {current_row - 3}")
    print(f"  Total columns: {current_col - col_offset - 1}")
    return output_file


def generate_analysis_sheet(wb, pivot, df_negative):
    """
    Generate a problem analysis sheet with multiple analysis sections.
    """
    import numpy as np
    
    ws = wb.create_sheet("Анализ проблем")
    current_row = 1
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    section_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    section_font = Font(bold=True, size=11)
    
    # SECTION 0: Most Frequent Subtopics
    current_row = write_section_header(ws, current_row, "САМЫЕ ЧАСТЫЕ ПОДТЕМЫ ЖАЛОБ", section_fill, section_font)
    current_row += 1
    
    headers = ["Ранг", "Подтема", "Тег", "Количество жалоб", "% от всех негативных"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
    current_row += 1
    
    # Group by subtopic and count
    subtopic_counts = pivot.groupby(['Sub_tag', 'Primary_Tag'])['Count'].sum().reset_index()
    subtopic_counts = subtopic_counts.sort_values('Count', ascending=False).head(20)
    
    total_negative = len(df_negative)
    
    for rank, (_, row) in enumerate(subtopic_counts.iterrows(), start=1):
        pct = (row['Count'] / total_negative * 100) if total_negative > 0 else 0
        ws.cell(row=current_row, column=1, value=rank)
        ws.cell(row=current_row, column=2, value=row['Sub_tag'])
        ws.cell(row=current_row, column=3, value=row['Primary_Tag'])
        ws.cell(row=current_row, column=4, value=int(row['Count']))
        ws.cell(row=current_row, column=5, value=f"{pct:.1f}%")
        current_row += 1
    
    current_row += 2
    
    # SECTION 1: Worst GOSB per Sub-topic
    current_row = write_section_header(ws, current_row, "ХУДШИЙ ГОСБ ПО КАЖДОЙ ПОДТЕМЕ", section_fill, section_font)
    current_row += 1
    
    headers = ["Подтема", "Тег", "ГОСБ", "Жалоб", "Среднее", "2-е место", "% от ГОСБ"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
    current_row += 1
    
    # For each subtopic, find GOSB with most complaints
    subtopic_gosb = pivot.groupby(['Sub_tag', 'Primary_Tag', 'GOSB'])['Count'].sum().reset_index()
    
    # Calculate stats for each subtopic
    subtopic_totals = subtopic_gosb.groupby('Sub_tag')['Count'].sum().to_dict()
    subtopic_means = subtopic_gosb.groupby('Sub_tag')['Count'].mean().to_dict()
    
    # Calculate total reviews per GOSB (from df_negative, not pivot)
    gosb_total_reviews = df_negative.groupby('GOSB').size().to_dict()
    
    results = []
    for subtag in subtopic_gosb['Sub_tag'].unique():
        subtag_data = subtopic_gosb[subtopic_gosb['Sub_tag'] == subtag].sort_values('Count', ascending=False)
        if len(subtag_data) > 0:
            worst = subtag_data.iloc[0]
            second_place = subtag_data.iloc[1]['Count'] if len(subtag_data) > 1 else 0
            avg = subtopic_means.get(subtag, 0)
            
            # Calculate % from total GOSB reviews
            gosb_total = gosb_total_reviews.get(worst['GOSB'], 1)
            pct = (worst['Count'] / gosb_total * 100) if gosb_total > 0 else 0
            
            results.append({
                'Sub_tag': worst['Sub_tag'],
                'Primary_Tag': worst['Primary_Tag'],
                'GOSB': worst['GOSB'],
                'Count': worst['Count'],
                'Avg': avg,
                'Second': second_place,
                'Pct': pct
            })
    
    results_df = pd.DataFrame(results).sort_values('Count', ascending=False)
    
    for _, row in results_df.iterrows():
        ws.cell(row=current_row, column=1, value=row['Sub_tag'])
        ws.cell(row=current_row, column=2, value=row['Primary_Tag'])
        ws.cell(row=current_row, column=3, value=row['GOSB'])
        ws.cell(row=current_row, column=4, value=int(row['Count']))
        ws.cell(row=current_row, column=5, value=f"{row['Avg']:.1f}")
        ws.cell(row=current_row, column=6, value=int(row['Second']))
        ws.cell(row=current_row, column=7, value=f"{row['Pct']:.1f}%")
        current_row += 1
    
    current_row += 2
    
    # SECTION 2: Worst Office per Sub-topic
    current_row = write_section_header(ws, current_row, "ХУДШЕЕ ОТДЕЛЕНИЕ ПО КАЖДОЙ ПОДТЕМЕ", section_fill, section_font)
    current_row += 1
    
    headers = ["Подтема", "Тег", "ГОСБ", "Отделение", "Жалоб", "Среднее", "2-е место", "% от отд."]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
    current_row += 1
    
    # For each subtopic, find office with most complaints
    # Calculate stats per subtopic
    subtopic_office_totals = pivot.groupby('Sub_tag')['Count'].sum().to_dict()
    subtopic_office_means = pivot.groupby('Sub_tag')['Count'].mean().to_dict()
    
    # Calculate total reviews per office (from df_negative, not pivot)
    office_total_reviews = df_negative.groupby('City_Address').size().to_dict()
    
    office_results = []
    for subtag in pivot['Sub_tag'].unique():
        subtag_data = pivot[pivot['Sub_tag'] == subtag].sort_values('Count', ascending=False)
        if len(subtag_data) > 0:
            worst = subtag_data.iloc[0]
            second_place = subtag_data.iloc[1]['Count'] if len(subtag_data) > 1 else 0
            avg = subtopic_office_means.get(subtag, 0)
            
            # Calculate % from total office reviews
            office_total = office_total_reviews.get(worst['City_Address'], 1)
            pct = (worst['Count'] / office_total * 100) if office_total > 0 else 0
            
            office_results.append({
                'Sub_tag': worst['Sub_tag'],
                'Primary_Tag': worst['Primary_Tag'],
                'GOSB': worst['GOSB'],
                'City_Address': worst['City_Address'],
                'Count': worst['Count'],
                'Avg': avg,
                'Second': second_place,
                'Pct': pct
            })
    
    office_results_df = pd.DataFrame(office_results).sort_values('Count', ascending=False)
    
    for _, row in office_results_df.iterrows():
        ws.cell(row=current_row, column=1, value=row['Sub_tag'])
        ws.cell(row=current_row, column=2, value=row['Primary_Tag'])
        ws.cell(row=current_row, column=3, value=row['GOSB'])
        ws.cell(row=current_row, column=4, value=row['City_Address'])
        ws.cell(row=current_row, column=5, value=int(row['Count']))
        ws.cell(row=current_row, column=6, value=f"{row['Avg']:.1f}")
        ws.cell(row=current_row, column=7, value=int(row['Second']))
        ws.cell(row=current_row, column=8, value=f"{row['Pct']:.1f}%")
        current_row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15

def write_section_header(ws, row, text, fill, font):
    """Helper to write and merge section headers"""
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = fill
    cell.font = font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    return row

if __name__ == "__main__":
    generate_heatmap()
